"""Builds a styled .xlsx version of the organized leads — colored header,
zebra striping, and the same invalid/review row highlighting shown in the
web table, plus a Score column. The plain CSV download stays exactly the
standard 10 columns; this is the "make it look nice to open in Excel" copy.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="5B4FCF")
_ZEBRA_FILL = PatternFill("solid", fgColor="F6F5FB")
_INVALID_FILL = PatternFill("solid", fgColor="FDECEB")
_REVIEW_FILL = PatternFill("solid", fgColor="FDF3E0")
_BORDER = Border(*(Side(style="thin", color="E6E4F2"),) * 4)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SCORE_COLORS = {"invalid": "D1453D", "review": "C9821A", "ok": "1A9C6B"}


def build_pretty_workbook(columns, rows, out_path):
    """rows: [{"cells": [...], "score": int, "review": bool, "invalid": bool}]"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Organized Leads"

    header = list(columns) + ["Score"]
    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = _BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    widths = [len(str(h)) for h in header]
    for i, row in enumerate(rows, start=2):
        values = list(row["cells"]) + [row["score"]]
        ws.append(values)
        status = "invalid" if row["invalid"] else ("review" if row["review"] else "ok")
        fill = {"invalid": _INVALID_FILL, "review": _REVIEW_FILL}.get(status)
        if fill is None and i % 2 == 0:
            fill = _ZEBRA_FILL
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx)
            cell.border = _BORDER
            if fill is not None:
                cell.fill = fill
            widths[col_idx - 1] = max(widths[col_idx - 1], len(str(value)))
        score_cell = ws.cell(row=i, column=len(values))
        score_cell.font = Font(bold=True, color=_SCORE_COLORS[status])
        score_cell.alignment = Alignment(horizontal="center")

    if len(rows):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 10), 42)

    wb.save(out_path)
